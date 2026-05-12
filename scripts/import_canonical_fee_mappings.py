import json
import re
from collections import OrderedDict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
CLIENT_SOURCE = Path(r"C:\Users\trevo\Wealthworks Dropbox\Wealth Works (Pty) Ltd\Roger Data\Client Names.xlsx")
FEE_SOURCE = Path(r"C:\Users\trevo\Wealthworks Dropbox\Wealth Works (Pty) Ltd\Roger Data\Wealthworks Fees 2026.xlsx")
OUT = ROOT / "base44" / "functions" / "_shared" / "canonicalFeeMappings.ts"

PROVIDERS = {
    "gryphon asset management": "Gryphon",
    "peresec securities": "Peresec",
    "northstar": "Northstar FNB",
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalise(value):
    value = clean(value).lower()
    value = re.sub(r"\b(mr|mrs|ms|miss|dr|prof)\b", " ", value)
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def compact(value):
    return re.sub(r"[^a-z0-9]+", "", normalise(value))


def provider_name(value):
    return PROVIDERS.get(normalise(value), clean(value))


def name_score(left, right):
    if compact(left) == compact(right):
        return 100

    left_parts = normalise(left).split()
    right_parts = normalise(right).split()
    if not left_parts or not right_parts or left_parts[0] != right_parts[0]:
        return 0

    left_initials = "".join(part[0] for part in left_parts[1:] if part)
    right_initials = "".join(part[0] for part in right_parts[1:] if part)
    if left_initials and right_initials and (left_initials.startswith(right_initials) or right_initials.startswith(left_initials)):
        return 90
    return 65


def investment_score(left, right):
    left_key = compact(left)
    right_key = compact(right)
    if not left_key or not right_key:
        return 0
    if left_key == right_key:
        return 100
    if left_key in right_key or right_key in left_key:
        return 88

    left_tokens = set(normalise(left).split())
    right_tokens = set(normalise(right).split())
    if not left_tokens or not right_tokens:
        return 0
    overlap = len(left_tokens & right_tokens)
    coverage = overlap / max(len(left_tokens), len(right_tokens))
    return 70 if coverage >= 0.6 else 0


def load_fee_rows():
    workbook = openpyxl.load_workbook(FEE_SOURCE, data_only=True)
    sheet = workbook["Q1 Assets"]
    rows = []
    for row_index in range(2, sheet.max_row + 1):
        client = clean(sheet.cell(row_index, 1).value)
        investment = clean(sheet.cell(row_index, 3).value)
        platform = provider_name(sheet.cell(row_index, 5).value)
        if not client or not investment or not platform:
            continue
        rows.append(
            {
                "client": client,
                "investmentName": investment,
                "platform": platform,
                "rebateAnnualPercent": float(sheet.cell(row_index, 9).value or 0) * 100,
                "advisoryAnnualPercent": float(sheet.cell(row_index, 10).value or 0) * 100,
            }
        )
    return rows


def load_client_rows():
    workbook = openpyxl.load_workbook(CLIENT_SOURCE, data_only=True)
    sheet = workbook.active
    rows = []
    for row_index in range(2, sheet.max_row + 1):
        account_code = clean(sheet.cell(row_index, 2).value)
        identity_no = clean(sheet.cell(row_index, 3).value)
        portfolio_name = clean(sheet.cell(row_index, 5).value)
        platform = provider_name(sheet.cell(row_index, 6).value)
        investment = clean(sheet.cell(row_index, 7).value)
        if not account_code or not portfolio_name or not platform or not investment:
            continue
        rows.append(
            {
                "accountCode": account_code,
                "identityNo": identity_no,
                "portfolioName": portfolio_name,
                "platform": platform,
                "investmentName": investment,
            }
        )
    return rows


def best_fee_match(client_row, fee_rows):
    candidates = []
    for fee_row in fee_rows:
        if compact(fee_row["platform"]) != compact(client_row["platform"]):
            continue
        n_score = name_score(client_row["portfolioName"], fee_row["client"])
        i_score = investment_score(client_row["investmentName"], fee_row["investmentName"])
        if n_score >= 65 and i_score >= 70:
            candidates.append((i_score * 2 + n_score, n_score, i_score, fee_row))
    return sorted(candidates, key=lambda item: item[0], reverse=True)[0] if candidates else None


def best_account_fallback(client_row, fee_rows):
    candidates = []
    for fee_row in fee_rows:
        if compact(fee_row["platform"]) != compact(client_row["platform"]):
            continue
        if compact(fee_row["investmentName"]) != "equities":
            continue
        n_score = name_score(client_row["portfolioName"], fee_row["client"])
        if n_score >= 65:
            candidates.append((n_score, fee_row))
    return sorted(candidates, key=lambda item: item[0], reverse=True)[0] if candidates else None


def main():
    fee_rows = load_fee_rows()
    client_rows = load_client_rows()
    exact_rules = OrderedDict()
    fallback_rules = OrderedDict()

    for client_row in client_rows:
        match = best_fee_match(client_row, fee_rows)
        if match:
            _, name_match, investment_match, fee_row = match
            key = (
                compact(client_row["platform"]),
                client_row["accountCode"].lower(),
                compact(client_row["investmentName"]),
            )
            exact_rules[key] = {
                **client_row,
                "rebateAnnualPercent": fee_row["rebateAnnualPercent"],
                "advisoryAnnualPercent": fee_row["advisoryAnnualPercent"],
                "sourceClient": fee_row["client"],
                "sourceInvestment": fee_row["investmentName"],
                "matchScore": name_match + investment_match,
            }
            continue

        fallback = best_account_fallback(client_row, fee_rows)
        if fallback:
            name_match, fee_row = fallback
            key = (compact(client_row["platform"]), client_row["accountCode"].lower())
            fallback_rules[key] = {
                "accountCode": client_row["accountCode"],
                "identityNo": client_row["identityNo"],
                "portfolioName": client_row["portfolioName"],
                "platform": client_row["platform"],
                "investmentName": "*",
                "rebateAnnualPercent": fee_row["rebateAnnualPercent"],
                "advisoryAnnualPercent": fee_row["advisoryAnnualPercent"],
                "sourceClient": fee_row["client"],
                "sourceInvestment": fee_row["investmentName"],
                "matchScore": name_match,
            }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        "export const canonicalFeeMappings = "
        + json.dumps(list(exact_rules.values()), indent=2)
        + ";\n\nexport const canonicalFeeFallbackMappings = "
        + json.dumps(list(fallback_rules.values()), indent=2)
        + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(exact_rules)} fee rules and {len(fallback_rules)} fallback rules to {OUT}")


if __name__ == "__main__":
    main()
