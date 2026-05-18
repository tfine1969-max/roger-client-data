"""Regenerates src/data/rogerSourceRows.js from the Roger Jan-Mar Excel file."""
import openpyxl, json, re, sys
from pathlib import Path

EXCEL = Path(r'C:\Users\trevo\Wealthworks Dropbox\Wealth Works (Pty) Ltd\Roger Data\Roger Data - Jan to March.xlsx')
OUT = Path(__file__).parent.parent / 'src' / 'data' / 'rogerSourceRows.js'

MONTH_MAP = {'Jan': '2026-01', 'Feb': '2026-02', 'March': '2026-03'}
MONTH_END = {'2026-01': '2026-01-31', '2026-02': '2026-02-28', '2026-03': '2026-03-31'}


def normalize(v):
    return re.sub(r'[^a-z0-9]+', ' ', str(v or '').strip().lower()).strip()


def clean(v):
    return str(v or '').strip()


def parse_num(v):
    if isinstance(v, (int, float)):
        return float(v) if v == v else None
    try:
        return float(str(v).replace(',', '').replace('R', '').strip())
    except Exception:
        return None


def to_platform(v):
    t = normalize(v)
    if 'gryphon' in t:
        return 'Gryphon'
    if 'julius' in t:
        return 'Julius Baer'
    if 'northstar' in t:
        return 'Northstar'
    if 'peresec' in t:
        return 'Peresec'
    if 'prescient' in t:
        return 'Prescient'
    if 'credo' in t:
        return 'Credo'
    if 'prime' in t:
        return 'Prime'
    return clean(v)


def account_code(plat, client):
    slug = re.sub(r'[^A-Z0-9]+', '_', client.upper()).strip('_')[:70]
    return f'{plat.upper()}_{slug or "UNKNOWN"}'


def find_col(headers, candidates):
    for i, h in enumerate(headers):
        for c in candidates:
            if h == c or (c and c in h):
                return i
    return None


wb = openpyxl.load_workbook(EXCEL, data_only=True)
rows = []
totals = {}

for sname in wb.sheetnames:
    upload_month = MONTH_MAP.get(sname)
    if not upload_month:
        continue
    ws = wb[sname]
    data = list(ws.iter_rows(values_only=True))

    hi = next((i for i, r in enumerate(data) if r and r[0] == 'Client'), None)
    if hi is None:
        print(f'WARNING: No header row found in {sname}', file=sys.stderr)
        continue

    headers = [normalize(h) for h in data[hi]]
    nav_col = find_col(headers, ['nav'])
    rebate_monthly_col = find_col(headers, ['rebates jan', 'rebates feb', 'rebates mar', 'rebates'])
    advisory_monthly_col = find_col(headers, ['advisory jan', 'advisory feb', 'advisory mar'])
    rebate_pct_col = find_col(headers, ['rebate'])
    advisory_pct_col = find_col(headers, ['advisory fee'])

    aum = 0.0
    sheet_rows = 0
    for row in data[hi + 1:]:
        if not row or row[0] is None:
            continue
        client_name = clean(row[0])
        inv = clean(row[5]) if len(row) > 5 else ''
        cls = clean(row[6]) if len(row) > 6 else ''
        prov = clean(row[7]) if len(row) > 7 else ''
        nav = parse_num(row[nav_col]) if nav_col is not None and len(row) > nav_col else None

        if not client_name or not inv or not prov or nav is None:
            continue

        def get_pct(col):
            v = parse_num(row[col]) if col is not None and len(row) > col else None
            v = v or 0.0
            return v * 100 if abs(v) <= 1 else v

        rebate_pct = get_pct(rebate_pct_col)
        advisory_pct = get_pct(advisory_pct_col)

        rebate_m = parse_num(row[rebate_monthly_col]) if rebate_monthly_col is not None and len(row) > rebate_monthly_col else None
        advisory_m = parse_num(row[advisory_monthly_col]) if advisory_monthly_col is not None and len(row) > advisory_monthly_col else None
        if rebate_m is None:
            rebate_m = nav * (rebate_pct / 100) / 12
        if advisory_m is None:
            advisory_m = nav * (advisory_pct / 100) / 12

        plat = to_platform(prov)
        aum += nav
        sheet_rows += 1
        rows.append({
            'upload_month': upload_month,
            'account_code': account_code(plat, client_name),
            'identity_no': None,
            'portfolio_name': client_name,
            'platform': plat,
            'source_provider': prov,
            'investment_name': inv,
            'investment_class': cls,
            'currency': 'ZAR',
            'month_end_market_value': nav,
            'original_currency_value': nav,
            'exchange_rate_to_zar': 1,
            'zar_value': nav,
            'exchange_rate_date': MONTH_END[upload_month],
            'exchange_rate_source': 'Roger source workbook',
            'conversion_status': 'ZAR Source Value',
            'rebate_fee_annual_percent': rebate_pct,
            'advisory_fee_annual_percent': advisory_pct,
            'rebate_fee_monthly_percent': round(rebate_pct / 12, 10),
            'advisory_fee_monthly_percent': round(advisory_pct / 12, 10),
            'rebate_fee_monthly_amount_original_currency': rebate_m,
            'advisory_fee_monthly_amount_original_currency': advisory_m,
            'rebate_fee_monthly_amount_zar': rebate_m,
            'advisory_fee_monthly_amount_zar': advisory_m,
            'total_monthly_fee_original_currency': rebate_m + advisory_m,
            'total_monthly_fee_zar': rebate_m + advisory_m,
            'fee_required': False,
            'fee_source': 'roger-source',
            'has_missing_account_code': False,
            'has_missing_identity_no': True,
            'has_missing_market_value': nav == 0,
            'has_unknown_value': False,
            'is_duplicate': False,
            'is_flagged': False,
        })

    totals[upload_month] = aum
    print(f'{sname}: {sheet_rows} rows, AUM = {aum:,.2f}')

print(f'Total rows: {len(rows)}')

rows_js = json.dumps(rows, indent=2, ensure_ascii=False)
totals_js = json.dumps(totals, indent=2)
content = f'// Generated from Roger Data - Jan to March.xlsx. ZAR NAV values are source-of-truth rows.\nexport const rogerSourceRows = {rows_js};\nexport const rogerSourceTotals = {totals_js};\n'

OUT.write_text(content, encoding='utf-8')
print(f'Written to {OUT}')
