import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\trevo\Wealthworks Dropbox\Website 2026\Website Design\Base44\Roger Data\1st quarter recon.xlsx")
CLIENT_NAMES_SOURCE = Path(r"C:\Users\trevo\Wealthworks Dropbox\Wealth Works (Pty) Ltd\Roger Data\Client Names.xlsx")
OUT = ROOT / "src" / "data" / "feeMapping.js"

NAV_COLUMNS = {
    "2026-01": 5,
    "2026-02": 6,
    "2026-03": 7,
}

PROVIDER_IDS = {
    "Credo": "Credo",
    "Gryphon Asset Management": "Gryphon",
    "Julius Baer": "Julius Baer",
    "Northstar": "Northstar FNB",
    "Peresec Securities": "Peresec",
    "Prescient": "Prescient",
    "Prime": "Prime",
}


def normalized(value):
    value = str(value or "").lower()
    value = re.sub(r"\b(mr|mrs|ms|miss|dr|prof)\b", " ", value)
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def compact(value):
    return re.sub(r"[^a-z0-9]+", "", normalized(value))


def client_tokens(value):
    norm = normalized(value)
    if not norm:
        return []
    parts = norm.split()
    return [part for part in parts if len(part) > 1]


def main():
    workbook = openpyxl.load_workbook(SOURCE, data_only=True)
    sheet = workbook.active
    client_names_workbook = openpyxl.load_workbook(CLIENT_NAMES_SOURCE, data_only=True)
    client_names_sheet = client_names_workbook.active
    indexed_client_names = [
        str(row[0]).strip()
        for row in client_names_sheet.iter_rows(min_row=2, values_only=True)
        if row and row[0]
    ]

    rows = []
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        client, _, investment, investment_class, service_provider = row[:5]
        rebate, advisory = row[8], row[9]
        if not client or not investment or not service_provider:
            continue
        corrected_client = indexed_client_names[index] if index < len(indexed_client_names) else str(client).strip()
        provider = PROVIDER_IDS.get(service_provider, service_provider)
        rows.append(
            {
                "client": corrected_client,
                "clientKey": compact(corrected_client),
                "clientTokens": client_tokens(corrected_client),
                "investment": str(investment).strip(),
                "investmentClass": str(investment_class or "").strip(),
                "investmentKey": compact(investment),
                "provider": provider,
                "sourceProvider": str(service_provider).strip(),
                "navByMonth": {
                    month: float(row[column] or 0)
                    for month, column in NAV_COLUMNS.items()
                },
                "rebateAnnualPercent": float(rebate or 0) * 100,
                "advisoryAnnualPercent": float(advisory or 0) * 100,
            }
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        "export const feeMappingRows = "
        + json.dumps(rows, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(rows)} fee mapping rows to {OUT}")


if __name__ == "__main__":
    main()
