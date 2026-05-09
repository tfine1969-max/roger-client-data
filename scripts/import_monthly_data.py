import json
from collections import defaultdict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = Path(r"C:\Users\trevo\Wealthworks Dropbox\Website 2026\Website Design\Base44\Roger Data")
OUT = ROOT / "src" / "data" / "monthlyClientData.js"

FILES = [
    ("jan-2026", "Jan 2026", SOURCE_DIR / "Jan.xlsx"),
    ("feb-2026", "Feb 2026", SOURCE_DIR / "feb.xlsx"),
]

PROVIDERS = {
    "Julius Baer": ("julius-baer", "Julius Baer"),
    "Credo": ("credo", "Credo"),
    "Gryphon": ("gryphon", "Gryphon"),
    "Prime": ("prime", "Prime Investments"),
    "Northstar FNB": ("northstar-fnb", "Northstar FNB"),
    "Northstar Sanlam": ("northstar-sanlam", "Northstar Sanlam"),
    "Peresec": ("peresec", "Peresec"),
    "Prescient": ("prescient", "Prescient"),
}

FEE_RATES = {
    "julius-baer": {
        "rebateAnnualRate": 0.003221070905091383,
        "advisoryAnnualRate": 0.0056674609242583695,
    },
    "credo": {"rebateAnnualRate": 0.003, "advisoryAnnualRate": 0.006},
    "gryphon": {"rebateAnnualRate": 0.0025, "advisoryAnnualRate": 0.005},
    "prime": {"rebateAnnualRate": 0.0025, "advisoryAnnualRate": 0.005},
    "northstar-fnb": {"rebateAnnualRate": 0.0, "advisoryAnnualRate": 0.005},
    "northstar-sanlam": {"rebateAnnualRate": 0.0, "advisoryAnnualRate": 0.005},
    "peresec": {"rebateAnnualRate": 0.0, "advisoryAnnualRate": 0.005},
    "prescient": {"rebateAnnualRate": 0.0, "advisoryAnnualRate": 0.005},
}


def fee(native_values, zar_aum, provider_id, fee_type):
    annual_rate = FEE_RATES[provider_id][f"{fee_type}AnnualRate"]
    native_fees = {code: value * annual_rate / 12 for code, value in native_values.items()}
    return {
        "annualRate": annual_rate,
        "nativeFees": native_fees,
        "zarFee": zar_aum * annual_rate / 12,
    }


def import_file(month_id, label, path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active
    usd_zar = float(sheet["M2"].value)
    exchange_rates = {"USD": usd_zar, "ZAR": 1}
    clients = {}
    source_native_totals = defaultdict(float)

    for row in range(2, sheet.max_row + 1):
        provider_name = sheet.cell(row, 5).value
        native_value = sheet.cell(row, 8).value
        if not provider_name or native_value is None:
            continue

        provider_id, normalized_provider = PROVIDERS.get(
            provider_name,
            (str(provider_name).lower().replace(" ", "-"), provider_name),
        )
        account_code = str(sheet.cell(row, 2).value or "")
        client_name = str(sheet.cell(row, 4).value or "Unmapped Client")
        currency = str(sheet.cell(row, 7).value or "ZAR")
        native_value = float(native_value)
        zar_value = native_value * exchange_rates.get(currency, 1)
        source_native_totals[currency] += native_value

        client_id = f"{provider_id}|{account_code}|{client_name}"
        if client_id not in clients:
            clients[client_id] = {
                "id": client_id,
                "client": client_name,
                "accountCode": account_code,
                "identityNo": str(sheet.cell(row, 3).value or ""),
                "providerId": provider_id,
                "providerName": normalized_provider,
                "nativeValues": defaultdict(float),
                "zarAum": 0,
                "holdingCount": 0,
                "holdings": [],
            }

        client = clients[client_id]
        client["nativeValues"][currency] += native_value
        client["zarAum"] += zar_value
        client["holdingCount"] += 1
        client["holdings"].append(
            {
                "investment": sheet.cell(row, 6).value or "Unmapped holding",
                "currency": currency,
                "nativeValue": native_value,
                "zarValue": zar_value,
            }
        )

    client_rows = []
    for client in clients.values():
        client["nativeValues"] = dict(sorted(client["nativeValues"].items()))
        rebate = fee(client["nativeValues"], client["zarAum"], client["providerId"], "rebate")
        advisory = fee(client["nativeValues"], client["zarAum"], client["providerId"], "advisory")
        total_native = defaultdict(float)
        for fee_row in (rebate, advisory):
            for code, value in fee_row["nativeFees"].items():
                total_native[code] += value
        client["fees"] = {
            "rebate": rebate,
            "advisory": advisory,
            "total": {"nativeFees": dict(total_native), "zarFee": rebate["zarFee"] + advisory["zarFee"]},
        }
        client_rows.append(client)

    source_native_totals = dict(sorted(source_native_totals.items()))
    source_zar_totals = {
        code: value * exchange_rates.get(code, 1)
        for code, value in source_native_totals.items()
    }

    return {
        "id": month_id,
        "label": label,
        "sourceFile": path.name,
        "exchangeRates": exchange_rates,
        "sourceNativeTotals": source_native_totals,
        "sourceZarTotals": source_zar_totals,
        "sourceZarTotal": sum(source_zar_totals.values()),
        "clients": sorted(client_rows, key=lambda item: (item["providerName"], item["client"], item["accountCode"])),
    }


def main():
    months = [import_file(*item) for item in FILES]
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        "export const feeRates = "
        + json.dumps(FEE_RATES, indent=2)
        + ";\n\nexport const monthlyClientData = "
        + json.dumps(months, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    for month in months:
        print(month["label"], f"{month['sourceZarTotal']:.2f}", month["sourceNativeTotals"])


if __name__ == "__main__":
    main()
